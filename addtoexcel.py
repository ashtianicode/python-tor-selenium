from openpyxl import Workbook
from openpyxl import load_workbook


# in the main file first we check os.path.isfile("dest_filename")
# if it not exsited call makeexcelfile

def makeexcelfile(dest_filename):
    excelfilename = dest_filename
    wb = Workbook()
    mobile = wb.create_sheet("sheetname")
    wb.save(dest_filename)


def addtoexcel(excelfilename,itemindex,phone):
            wb = load_workbook(excelfilename)
            ws_text = wb["sheetname"]
            lastrow = len(tuple(ws_text.rows)) 
            ws_text.cell(row=lastrow +1, column=1, value="____")
            wb.save(filename = excelfilename)
